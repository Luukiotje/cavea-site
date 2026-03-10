"""
=============================================================
  CAVEA BLOG AUTOMATION — MAIN RUNNER (v3)
  This is the script you run to publish one blog post.
  It combines the generator and publisher into one command.

  CHANGES in v3:
  - Passes the tag from Claude to the publisher (no more hardcoded)
  - Updates Excel tracking file with USED status
  - Commits used_images.json alongside topics.json
=============================================================

USAGE:
  # Publish the next scheduled post:
  python scripts/run_automation.py

  # Publish a specific topic by ID (for manual extra posts):
  python scripts/run_automation.py --id 5

  # Preview without actually publishing (dry run):
  python scripts/run_automation.py --dry-run
"""

import sys
import os
import argparse
import json
import datetime

# Add parent directory so we can import our scripts
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(BASE_DIR, "scripts"))

from blog_generator import (
    get_next_topic, generate_blog_content, parse_claude_response,
    fetch_wine_image, build_html_page, mark_topic_published,
    TOPICS_PATH, POSTS_DIR
)
from github_publisher import publish_post


def get_topic_by_id(topic_id):
    with open(TOPICS_PATH) as f:
        topics = json.load(f)
    for topic in topics:
        if topic["id"] == topic_id:
            return topic, topics
    return None, topics


def update_excel_tracking():
    """
    Updates the Excel tracking file (cavea_automation/blog_tracking.xlsx)
    with current topic statuses. Each topic gets a row with a "USED" marker
    when published, plus the publication date.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        xlsx_path = os.path.join(BASE_DIR, "blog_tracking.xlsx")

        # Load current topics
        with open(TOPICS_PATH) as f:
            topics = json.load(f)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Blog Topics"

        # Styling
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
        gold_font = Font(color="C4A24E", bold=True)
        used_fill = PatternFill(start_color="1A3D1A", end_color="1A3D1A", fill_type="solid")
        pending_fill = PatternFill(start_color="3D3D1A", end_color="3D3D1A", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='555555'),
            right=Side(style='thin', color='555555'),
            top=Side(style='thin', color='555555'),
            bottom=Side(style='thin', color='555555')
        )

        # Headers
        headers = ["ID", "Onderwerp", "Titel", "Slug", "STATUS", "Publicatiedatum", "Keyword Focus"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Data rows
        for row_idx, topic in enumerate(topics, 2):
            is_published = topic.get("status") == "published"
            status_text = "USED" if is_published else "PENDING"
            pub_date = topic.get("published_date", "")

            values = [
                topic["id"],
                topic["short_tail"],
                topic["title_1"],
                topic["slug"],
                status_text,
                pub_date,
                topic["keyword_focus_1"]
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border

                # Style the STATUS column
                if col == 5:
                    cell.font = gold_font
                    cell.fill = used_fill if is_published else pending_fill
                    cell.alignment = Alignment(horizontal='center')

        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 55
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 30

        wb.save(xlsx_path)
        print(f"  Excel tracking bijgewerkt: {xlsx_path}")
        return True

    except ImportError:
        print("  openpyxl niet beschikbaar, Excel tracking overgeslagen")
        return False
    except Exception as e:
        print(f"  Excel tracking fout: {e}")
        return False


def run(topic_id=None, dry_run=False):
    print("\n" + "="*55)
    print("  🍷 CAVEA BLOG AUTOMATION")
    print("="*55)

    # 1. Get topic
    if topic_id:
        topic, all_topics = get_topic_by_id(topic_id)
        if not topic:
            print(f"❌ Onderwerp met ID {topic_id} niet gevonden.")
            return
    else:
        topic, all_topics = get_next_topic()
        if not topic:
            print("✅ Alle geplande blogposts zijn al gepubliceerd!")
            return

    print(f"\n📋 Onderwerp: {topic['title_1']}")
    print(f"   Slug:      {topic['slug']}")
    print(f"   Status:    {topic.get('status', 'pending')}")

    # 2. Fetch a UNIQUE image (no duplicates across posts)
    print("\n🖼️  Zoekt unieke relevante foto...")
    image = fetch_wine_image(topic['short_tail'])
    print(f"   Foto URL: {image['url'][:60]}...")

    # 3. Generate content with Claude AI
    print("\n✍️  Claude schrijft het artikel...")
    raw_content = generate_blog_content(topic)
    parsed = parse_claude_response(raw_content)

    if not parsed["content"]:
        print("❌ Geen inhoud gegenereerd. Probeer opnieuw.")
        return

    print(f"   Meta description: {parsed['meta_description'][:60]}...")
    print(f"   Categorie: {parsed['tag']}")

    # 4. Build HTML page
    publish_date = datetime.date.today().isoformat()
    html_content = build_html_page(topic, parsed, image, publish_date)

    # 5. Save locally
    local_path = os.path.join(POSTS_DIR, f"{topic['slug']}.html")
    with open(local_path, "w", encoding="utf-8") as f:
        f.write(html_content)
    print(f"\n💾 Lokaal opgeslagen: {local_path}")

    if dry_run:
        print("\n🔍 DRY RUN — niet gepubliceerd naar GitHub.")
        print("   Bekijk het bestand lokaal om te controleren.")
        return

    # 6. Publish to GitHub (pass the tag from Claude)
    live_url = publish_post(
        topic, html_content, image["url"],
        parsed["meta_description"], tag=parsed["tag"]
    )

    # 7. Mark as published in topics.json
    mark_topic_published(all_topics, topic["id"])

    # 8. Update Excel tracking file
    update_excel_tracking()

    if live_url:
        print(f"\n🎉 KLAAR! Blogpost live op:")
        print(f"   {live_url}")
    else:
        print("\n⚠️  Publicatie mislukt. Controleer je GitHub token en repository naam.")

    print("\n" + "="*55 + "\n")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Cavea Blog Automation")
    parser.add_argument("--id",      type=int, help="Publiceer specifiek onderwerp op ID")
    parser.add_argument("--dry-run", action="store_true", help="Genereer maar publiceer niet")
    args = parser.parse_args()

    run(topic_id=args.id, dry_run=args.dry_run)
